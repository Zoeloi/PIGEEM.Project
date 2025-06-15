import openpyxl 
from openpyxl import Workbook, load_workbook
from datetime import datetime 
import os 

def menu_finanzas():
    while True:
        print("1) Ver Egresos")
        print("2) Ver Ingresos")    
        print("3) Situacion Financiera")
        print("4) Salir")
        opcion = input("Seleccione una opción: ")
        if opcion == "1":
            return egresos()
        elif opcion == "2":
            print("Xd")
        elif opcion == "3":
            print("Mostrando Situación Financiera...")
        elif opcion == "4":
            return
        else:
            print("Opción no válida. Intente de nuevo.")


def egresos():
    while True:
        print("\n--- Menú de Egresos ---")
        print("1) Ver reporte")
        print("2) Ingresar Gastos")
        print("3) Volver al menú principal")
        opcion1 = input("Seleccione una opción: ")


        if opcion1 == "1":
             #lógica para mostrar el reporte
            print("¿Qué reporte desea ver?")
            print("1) Egresos del día")
            print("2) Egresos de la semana")
            print("3) Egresos del mes")
            filtro_op = input("Seleccione una opción: ")
            if filtro_op == "1":
                filtro = "dia"
            elif filtro_op == "2":
                filtro = "semana"
            elif filtro_op == "3":
                filtro = "mes"
            else:
                print("Opción no válida.")
                return

            EXCEL_FILE = "egresos.xlsx"
            if not os.path.exists(EXCEL_FILE):
                print("No hay egresos registrados.")
                return
            wb = load_workbook(EXCEL_FILE)
            ws = wb.active
            hoy = datetime.now().date()
            print("\n--- Reporte de Egresos ---")
            for row in ws.iter_rows(min_row=2, values_only=True):
                fecha, monto, descripcion = row
                fecha_dt = datetime.strptime(fecha, "%Y-%m-%d").date()
                if filtro == "dia" and fecha_dt == hoy:
                    print(f"{fecha} | ${monto} | {descripcion}")
                elif filtro == "semana" and (hoy - fecha_dt).days < 7:
                    print(f"{fecha} | ${monto} | {descripcion}")
                elif filtro == "mes" and fecha_dt.month == hoy.month and fecha_dt.year == hoy.year:
                    print(f"{fecha} | ${monto} | {descripcion}")

        elif opcion1 == "2":
            print("Ingresando nuevo gasto...")
            monto = input("Monto del gasto: ")
            descripcion = input("Descripción del gasto: ")
            fecha = datetime.now().date()
            EXCEL_FILE = "egresos.xlsx"
            if not os.path.exists(EXCEL_FILE):
                wb = Workbook()
                ws = wb.active
                ws.append(["Fecha", "Monto", "Descripción"])
            else:
                wb = load_workbook(EXCEL_FILE)
                ws = wb.active
            ws.append([str(fecha), monto, descripcion])
            wb.save(EXCEL_FILE)
            print("Gasto registrado correctamente.")

        elif opcion1 == "3":
            print("Volviendo al menú principal...")
            return 
        else:
            print("Opción no válida. Intente de nuevo.")

def ingresos():
    while True:
        print("\n--- Menú de Ingresos ---")
        print("1) Ver reporte")
        print("2) Ingresar Ingresos ")
        print("3) Volver al menú principal")
        opcion2 = input("Seleccione una opción: ")

        if opcion1 == "1":
             #lógica para mostrar el reporte
            print("¿Qué reporte desea ver?")
            print("1) Egresos del día")
            print("2) Egresos de la semana")
            print("3) Egresos del mes")
            filtro_op = input("Seleccione una opción: ")
            if filtro_op == "1":
                filtro = "dia"
            elif filtro_op == "2":
                filtro = "semana"
            elif filtro_op == "3":
                filtro = "mes"
            else:
                print("Opción no válida.")
                return

            EXCEL_FILE = "egresos.xlsx"
            if not os.path.exists(EXCEL_FILE):
                print("No hay egresos registrados.")
                return
            wb = load_workbook(EXCEL_FILE)
            ws = wb.active
            hoy = datetime.now().date()
            print("\n--- Reporte de Egresos ---")
            for row in ws.iter_rows(min_row=2, values_only=True):
                fecha, monto, descripcion = row
                fecha_dt = datetime.strptime(fecha, "%Y-%m-%d").date()
                if filtro == "dia" and fecha_dt == hoy:
                    print(f"{fecha} | ${monto} | {descripcion}")
                elif filtro == "semana" and (hoy - fecha_dt).days < 7:
                    print(f"{fecha} | ${monto} | {descripcion}")
                elif filtro == "mes" and fecha_dt.month == hoy.month and fecha_dt.year == hoy.year:
                    print(f"{fecha} | ${monto} | {descripcion}")



    






if __name__ == "__main__":
    menu_finanzas()
